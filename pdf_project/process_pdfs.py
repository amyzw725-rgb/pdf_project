import subprocess
import pdfplumber
import pandas as pd
import os
import re
from pathlib import Path
import pytesseract
import shutil
from pdf2image import convert_from_path
from datetime import datetime, timedelta
import logging
import cv2
import numpy as np
from collections import Counter
from typing import Callable, Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from poppler_setup import resolve_poppler_bin

# -----------------------------
# 配置路径（随脚本所在项目目录迁移；可用环境变量覆盖）
#
# 环境变量（可选）:
#   TESSERACT_CMD      Tesseract 可执行文件路径；未设置时优先用项目内
#                      Tesseract-OCR/tesseract.exe，否则用 Program Files 默认安装路径
#   POPPLER_PATH       Poppler 的 bin 目录（需含 pdfinfo.exe）；未设置时自动探测
#                      <项目根>/poppler/... 或 PATH；若无则尝试下载到 <项目根>/poppler
#   INVOICE_AUTO_INSTALL_POPPLER  设为 0/false/off 可禁止自动下载 Poppler
#   INVOICE_INPUT_DIR  输入 PDF 目录，默认 <项目根>/input
#   INVOICE_OUTPUT_DIR 输出 Excel 目录，默认 <项目根>/output
#   INVOICE_ARCHIVE_DIR 归档目录，默认 <项目根>/archive
#
# 项目根目录 _PROJECT_ROOT = 本文件 (process_pdfs.py) 所在文件夹，与当前工作目录无关。
# -----------------------------
_PROJECT_ROOT = Path(__file__).resolve().parent

if not logging.root.handlers:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

_tess = (os.environ.get("TESSERACT_CMD") or "").strip()
if not _tess:
    _bundled = _PROJECT_ROOT / "Tesseract-OCR" / "tesseract.exe"
    _tess = str(_bundled) if _bundled.is_file() else r"C:\Program Files\Tesseract-OCR\tesseract.exe"
pytesseract.pytesseract.tesseract_cmd = _tess

POPPLER_PATH = resolve_poppler_bin(_PROJECT_ROOT)

input_folder = (os.environ.get("INVOICE_INPUT_DIR") or "").strip() or str(_PROJECT_ROOT / "input")
output_folder = (os.environ.get("INVOICE_OUTPUT_DIR") or "").strip() or str(_PROJECT_ROOT / "output")
archive_root_folder = (os.environ.get("INVOICE_ARCHIVE_DIR") or "").strip() or str(_PROJECT_ROOT / "archive")
debug_target_pdf = "202604030017.pdf"
COMPANY_SHORT_NAME_ALIASES = {
    "comets": "Comets",
    "comets international": "Comets",
    "comets international limited": "Comets",
    "a plus japan": "APlusJapan",
    "aplus japan": "APlusJapan",
    "a plus japan inc": "APlusJapan",
    "aplus japan inc": "APlusJapan",
}

today = datetime.now().strftime("%Y%m%d")
output_file = os.path.join(output_folder, f"invoice_result_{today}.xlsx")

# 如果输出文件夹不存在，则创建
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

_tess_path = Path(pytesseract.pytesseract.tesseract_cmd)
_poppler_pdfinfo = Path(POPPLER_PATH) / "pdfinfo.exe"
logging.info(
    "配置: PROJECT_ROOT=%s",
    _PROJECT_ROOT,
)
logging.info(
    "路径: input=%s | output=%s | archive=%s",
    input_folder,
    output_folder,
    archive_root_folder,
)
logging.info(
    "工具: Tesseract=%s (存在=%s) | Poppler=%s | pdfinfo.exe (存在=%s)",
    pytesseract.pytesseract.tesseract_cmd,
    _tess_path.is_file(),
    POPPLER_PATH,
    _poppler_pdfinfo.is_file(),
)


def check_poppler():
    """Run a quick poppler check and log warnings if unavailable."""
    try:
        subprocess.run(
            [os.path.join(POPPLER_PATH, "pdfinfo.exe"), "-v"],
            check=True,
            capture_output=True,
            text=True,
        )
    except Exception as e:
        logging.warning(f"Poppler检查失败，PDF转图片可能不可用: {e}")

# -----------------------------
# 常用正则模式
# -----------------------------
CURRENCY_PATTERN = r'(USD|EUR|GBP|JPY|CNY|RMB|HKD|SGD|\$|€|£|¥|￥)'
TERM_PATTERN = r'(?:Net|Due|Payment)\s+(\d+)\s*(?:days?)?'

def extract_pdf_text(file_path):
    text = ""
    used_ocr = False

    # 1️⃣ 先尝试直接读文本
    try:
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text += page.extract_text() or ""
    except Exception as e:
        logging.warning(f"pdfplumber失败: {e}")

    # 2️⃣ 如果文本太少，再OCR
    if len(text.strip()) < 50:
        used_ocr = True
        logging.info(f"启用OCR: {file_path}")
        try:
            text = _ocr_pdf_to_text(file_path)
        except Exception as e:
            logging.error(f"OCR失败: {e}")
            return "", used_ocr

    logging.info(f"提取文本长度: {len(text)}")
    return text, used_ocr


def preprocess_image(pil_img):
    # PIL → numpy
    img = np.array(pil_img)

    # 灰度
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # 放大（非常重要）
    gray = cv2.resize(gray, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)

    # 去噪
    gray = cv2.GaussianBlur(gray, (3, 3), 0)

    # 去边框噪声（在阈值化之前处理，使 Otsu 与 OCR 输入一致）
    gray = cv2.copyMakeBorder(gray, 10, 10, 10, 10, cv2.BORDER_CONSTANT, value=255)

    # 二值化（关键）
    _, thresh = cv2.threshold(
        gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU
    )

    return thresh


def _ocr_pdf_to_text(file_path: str) -> str:
    """
    Poppler (pdf2image) rasterize → preprocess_image → Tesseract.
    Used when pdfplumber yields little text.
    """
    images = convert_from_path(
        file_path,
        poppler_path=POPPLER_PATH,
        dpi=300,
    )
    configs = [
        "--oem 3 --psm 6",
        "--oem 3 --psm 4",
        "--oem 3 --psm 11",
    ]
    ocr_text = ""
    for img in images:
        processed_img = preprocess_image(img)
        for cfg in configs:
            ocr_text += (
                pytesseract.image_to_string(
                    processed_img,
                    lang="eng+chi_sim",
                    config=cfg,
                )
                + "\n"
            )
    return ocr_text


def extract_seller_name(text, debug: bool = False, debug_filename: str = ""):

    # 关键字行标识
    SELLER_KEYWORDS = [
        "Seller", "From", "公司", "Supplier", "Vendor",
        "Issued by", "Account Name", "Payee", "Recipient", "户名", "Co., Limited"
    ]
    COMPANY_SUFFIX_RE = re.compile(
        r"\b(Co\.?,?\s*Limited|Limited|Ltd\.?|Inc\.?|LLC|Company)\b|有限公司",
        re.I,
    )

    def _normalize_company_for_exclusion(s: str) -> str:
        # Lowercase and remove spaces/punctuation so variants match:
        # "APLUS JAPAN, Inc" == "APLUSJAPAN,INC" == "Aplus Japan Inc"
        return re.sub(r"[^a-z0-9]+", "", (s or "").lower())

    # Exclude our own company name (case-insensitive, punctuation/space-insensitive)
    EXCLUDED_COMPANY_SIGNATURES = {
        "aplusjapaninc",
        "aplusjapan",  # also exclude if "Inc" is missing in the text
    }

    # 文本按行清理空行
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    candidates = []
    keyword_adjacent_candidates = []
    if debug:
        logging.info(
            f"[company-extract][{debug_filename}] total_nonempty_lines={len(lines)}; top_preview={lines[:10]}"
        )

    # 1️⃣ 关键字行下方1~3行作为候选
    for i, line in enumerate(lines):
        for kw in SELLER_KEYWORDS:
            if kw.lower() in line.lower():
                for offset in range(1, 4):
                    if i + offset < len(lines):
                        v = lines[i + offset]
                        candidates.append(v)
                        keyword_adjacent_candidates.append(v)

    # 2️⃣ 顶部前10行加入候选
    candidates += lines[:10]
    if debug:
        logging.info(
            f"[company-extract][{debug_filename}] candidates_count={len(candidates)}; candidates_preview={candidates[:20]}"
        )

    # 3️⃣ 清理前缀，并清理行尾 "Invoice No..." 之类的干扰字段
    cleaned = []
    for c in candidates:
        # Note: "Bill To"/"Payer" usually refers to the buyer, not the seller.
        c_clean = re.sub(r'^(Beneficiary|Payee|From|To|Bill\s*To|Payer)[:：]\s*', '', c, flags=re.I).strip()
        c_clean = re.sub(r'\s+Invoice\s*(No\.?|Number)?[:：].*$', '', c_clean, flags=re.I).strip()
        if len(c_clean) > 2:
            if re.search(r"\bbill\s*to\b|\bpayer\b", c_clean, re.I):
                continue
            sig = _normalize_company_for_exclusion(c_clean)
            if sig in EXCLUDED_COMPANY_SIGNATURES:
                if debug:
                    logging.info(
                        f"[company-extract][{debug_filename}] excluded_candidate='{c_clean}' signature='{sig}'"
                    )
                continue
            cleaned.append(c_clean)

    if debug:
        logging.info(
            f"[company-extract][{debug_filename}] cleaned_count={len(cleaned)}; cleaned_preview={cleaned[:20]}"
        )

    # 4️⃣ 优先匹配带公司后缀的候选（如 Co., Limited / Inc / LLC）
    suffix_candidates = []
    for c in cleaned:
        if COMPANY_SUFFIX_RE.search(c):
            if not re.search(r'Bank|Payment|Swift|IBAN|Account', c, re.I):
                suffix_candidates.append(c)
    if suffix_candidates:
        best = min(suffix_candidates, key=len)
        if debug:
            logging.info(
                f"[company-extract][{debug_filename}] selected_from_suffix_candidates='{best}'"
            )
        return best

    # 5️⃣ 过滤合理公司名（优先：包含公司后缀）
    company_keywords = ["Ltd", "LLC", "Inc", "Co.", "公司", "有限公司", "Limited"]
    filtered = []
    for c in cleaned:
        if any(k.lower() in c.lower() for k in company_keywords):
            # Keep valid company lines even if they originally had "Invoice No" (already stripped above).
            if not re.search(r'Bank|Payment|Swift|IBAN|Account', c, re.I):
                filtered.append(c)

    if debug:
        logging.info(
            f"[company-extract][{debug_filename}] filtered_count={len(filtered)}; filtered_preview={filtered[:20]}"
        )

    # 6️⃣ 出现次数最多的候选（公司）
    if filtered:
        for name, _count in Counter(filtered).most_common():
            if _normalize_company_for_exclusion(name) not in EXCLUDED_COMPANY_SIGNATURES:
                if debug:
                    logging.info(
                        f"[company-extract][{debug_filename}] selected_from_filtered='{name}'"
                    )
                return name
            elif debug:
                sig = _normalize_company_for_exclusion(name)
                logging.info(
                    f"[company-extract][{debug_filename}] skipped_filtered='{name}' signature='{sig}'"
                )

    # 7️⃣ fallback: 前10行含字母的行
    for line in lines[:10]:
        line_clean = re.sub(r'^(Beneficiary|Payee|From|To|Bill\s*To|Payer)[:：]\s*', '', line, flags=re.I).strip()
        line_clean = re.sub(r'\s+Invoice\s*(No\.?|Number)?[:：].*$', '', line_clean, flags=re.I).strip()
        if len(line_clean) > 2:
            if re.search(r"\bbill\s*to\b|\bpayer\b", line_clean, re.I):
                continue
            sig = _normalize_company_for_exclusion(line_clean)
            if sig in EXCLUDED_COMPANY_SIGNATURES:
                if debug:
                    logging.info(
                        f"[company-extract][{debug_filename}] excluded_fallback_line='{line_clean}' signature='{sig}'"
                    )
                continue
            if any(c.isalpha() for c in line_clean):
                if re.search(r'earliest due date|consolidated bills|framework', line_clean, re.I):
                    continue
                if debug:
                    logging.info(
                        f"[company-extract][{debug_filename}] selected_from_fallback='{line_clean}'"
                    )
                return line_clean

    # 8️⃣ KOL / individual fallback (no Inc/Ltd/etc. present)
    # Prefer keyword-adjacent lines and "name/handle-like" strings.
    def _is_kol_like(s: str) -> bool:
        if not s or len(s) < 3:
            return False
        if re.search(r'Bank|Payment|Invoice|Total|Amount|Date', s, re.I):
            return False
        if any(ch.isdigit() for ch in s):
            return False
        # allow letters, spaces and common name/handle punctuation
        return bool(re.fullmatch(r"[A-Za-z][A-Za-z\s\-\.'’&/]*[A-Za-z]$", s.strip()))

    kol_pool = []
    for s in keyword_adjacent_candidates + cleaned:
        s2 = s.strip()
        if _normalize_company_for_exclusion(s2) in EXCLUDED_COMPANY_SIGNATURES:
            continue
        if _is_kol_like(s2):
            kol_pool.append(s2)

    if kol_pool:
        # Rank: keyword-adjacent first, then shorter/cleaner.
        keyword_set = {v.strip() for v in keyword_adjacent_candidates}
        best = sorted(
            kol_pool,
            key=lambda x: (
                0 if x in keyword_set else 1,
                len(x),
            ),
        )[0]
        if debug:
            logging.info(
                f"[company-extract][{debug_filename}] selected_kol_like='{best}' pool_preview={kol_pool[:10]}"
            )
        return best

    if debug:
        logging.info(f"[company-extract][{debug_filename}] result=未识别")
    return "未识别"


def extract_currency_amount(text):
    """
    提取货币和金额：
    1️⃣ 先做一次全局货币扫描（第一轮筛选）
    2️⃣ 再在含有 total/amount 等关键词的行里优先提取金额
    3️⃣ 如果还不行，再从全文中找最大金额
    """
    # 1️⃣ 先做“第一轮筛选”：全局找货币符号/代码
    currency_match = re.search(CURRENCY_PATTERN, text)
    currency = currency_match.group(0) if currency_match else ""

    def _safe_parse_amount(raw: str):
        """Parse amount text safely; return None for malformed values."""
        if not raw:
            return None
        cleaned = raw.replace(",", "").strip()
        if not cleaned or not re.fullmatch(r"\d+(?:\.\d{2})?", cleaned):
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None

    # 2️⃣ 在包含 total/amount 等关键词的行中找金额
    total_keywords = ['total', 'amount', 'grand total', '合计']
    lines = text.split('\n')

    for line in lines:
        if any(kw.lower() in line.lower() for kw in total_keywords):
            # Must start with a digit, so punctuation-only tokens (e.g. ",") are ignored.
            match = re.search(r'\d[\d,]*(?:\.\d{2})?', line)
            if match:
                amount = _safe_parse_amount(match.group())
                if amount is None:
                    continue
                return currency, amount

    # 3️⃣ fallback：直接抓全文中的最大金额
    amounts = re.findall(r'\d[\d,]*\.\d{2}', text)
    if amounts:
        values = []
        for a in amounts:
            parsed = _safe_parse_amount(a)
            if parsed is not None:
                values.append(parsed)
        if not values:
            return currency, None
        return currency, max(values)

    return currency, None

def parse_date(date_str):
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y", "%d/%m/%Y", "%m-%d-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    try:
        return datetime.strptime(date_str, "%d %b %Y")
    except ValueError:
        return None


def extract_due_date(text):
    due_match = re.search(r'(Due Date|Payment Due|到期日)[^\n]{0,30}?(\d{4}[-/]\d{1,2}[-/]\d{1,2})', text, re.I)
    if due_match:
        parsed_date = parse_date(due_match.group(2))
        if parsed_date:
            return parsed_date.strftime("%Y-%m-%d")
    invoice_match = re.search(r'(Invoice Date|Date|发票日期)[^\n]{0,30}?(\d{4}[-/]\d{1,2}[-/]\d{1,2})', text, re.I)
    invoice_date = parse_date(invoice_match.group(2)) if invoice_match else None
    if invoice_date:
        term_match = re.search(TERM_PATTERN, text, re.I)
        days = int(term_match.group(1)) if term_match else 30
        due_date = invoice_date + timedelta(days=days)
        return due_date.strftime("%Y-%m-%d")
    return "未识别"

def extract_project_name(text, filename: str = ""):
    lines = text.split('\n')
    combined = f"{filename}\n{text}".lower()
    combined_no_space = re.sub(r"\s+", "", combined)

    # 1) Strong normalized project aliases (check filename + content)
    if "bleach" in combined or "bleach" in combined_no_space:
        return "Bleach"

    bpsr_aliases = [
        "bpsr",
        "blue protocol",
        "blue potocal",
        "blue portocal",
        "blue protocol star resonance",
        "star resonance",
        "星痕共鸣",
    ]
    if any(alias in combined for alias in bpsr_aliases):
        return "BPSR"
    # OCR may insert spaces between CJK chars, e.g. "星 痕 共 鸣".
    alias_no_space = [re.sub(r"\s+", "", alias.lower()) for alias in bpsr_aliases]
    if any(alias in combined_no_space for alias in alias_no_space):
        return "BPSR"

    # 2) Try structured "project/campaign" lines, but avoid pure invoice/meta noise
    noise_re = re.compile(
        r"(invoice\s*(no|number)?|bank|swift|iban|account|payer|payee|date|amount|total|thank you|sincerely|working with us|efforts)",
        re.I,
    )
    for line in lines:
        l = line.lower().strip()
        if any(k in l for k in ["project", "campaign"]):
            candidate = line.strip()
            # Reject paragraph-like sentences that include generic wording but not a project label.
            is_sentence_like = (("." in candidate) and (len(candidate) > 60))
            if candidate and not noise_re.search(candidate) and not is_sentence_like:
                return candidate

    return "未识别"

def extract_seller_address(text):
    address_keywords = ['Address', 'Addr', '地址']
    lines = text.split('\n')
    for i, line in enumerate(lines):
        for kw in address_keywords:
            if kw.lower() in line.lower():
                addr_lines = []
                for j in range(i, min(i+5, len(lines))):
                    if not lines[j].strip():
                        break
                    addr_lines.append(lines[j].strip())
                return ' '.join(addr_lines)
    return ""

def classify_country(company_name, seller_address, currency):
    text = (company_name or "") + " " + (seller_address or "")

    # 🇯🇵 日本 / 日本公司
    if "株式会社" in text:
        return "海外"

    if "Inc" in text or "LLC" in text:
        return "海外"

    if currency in ['USD', 'EUR', 'GBP', 'JPY']:
        return "海外"

    if currency in ['CNY', 'RMB', '￥']:
        return "国内"

    return "未知"

def classify_country_by_currency(currency):
    """根据货币判断国内/海外"""
    if not currency:
        return "未知"

    currency = currency.upper()
    domestic = {"CNY", "RMB", "￥", "元", "人民币"}
    overseas = {"USD", "US$", "EUR", "JPY", "HKD", "AUD", "GBP", "CAD", "SGD"}

    if currency in domestic:
        return "国内"
    elif currency in overseas:
        return "海外"
    else:
        return "未知"



def normalize_currency(curr):
    """统一货币格式，支持半角和全角符号"""
    if not curr:
        return "未知"

    # 去掉空格、换行、不可见字符
    curr = curr.strip().replace("\n", "").replace("\r", "").replace(" ", "")

    # 全角转半角（$、€、¥）
    full_to_half = {
        "＄": "$",
        "￥": "¥",
        "€": "€",
        "£": "£",
    }
    curr = full_to_half.get(curr, curr)

    # 转大写
    curr = curr.upper()

    # 映射到标准货币
    mapping = {
        "$": "USD",
        "€": "EUR",
        "£": "GBP",
        "¥": "CNY",
        "RMB": "CNY",
        "HKD": "HKD",
        "USD": "USD",
        "EUR": "EUR",
        "GBP": "GBP",
        "JPY": "JPY",
        "SGD": "SGD",
        "CNY": "CNY",
    }
    return mapping.get(curr, curr)

def normalize_company_name(name: str) -> str:
    """
    Convert extracted company name to Title Case by capitalizing the first
    letter of each word-like (alphabetic) segment while preserving punctuation.
    """
    if not name or not isinstance(name, str):
        return name

    # Remove leading OCR artifacts, e.g. "¢) ", "• ", ") ", etc.
    name = re.sub(r"^[^A-Za-z\u4e00-\u9fff0-9]+(?:\s*[\)\]\}])?\s*", "", name).strip()

    if not any(c.isalpha() for c in name):
        return name

    return re.sub(
        r"[A-Za-z]+",
        lambda m: m.group(0)[:1].upper() + m.group(0)[1:].lower(),
        name,
    )

def extract_fields(text, filename):

    debug_company = (filename == "202601_BPSR_Braxophone_KOL_Invoice.pdf")
    raw_seller_name = extract_seller_name(text, debug=debug_company, debug_filename=filename)
    seller_name = normalize_company_name(raw_seller_name)
    if debug_company:
        logging.info(
            f"[company-extract][{filename}] raw_extracted='{raw_seller_name}' normalized_title_case='{seller_name}'"
        )
    seller_address = extract_seller_address(text)
    currency, amount = extract_currency_amount(text)
    currency = normalize_currency(currency)
    payment_date = extract_due_date(text)
    project_name = extract_project_name(text, filename)

    country = classify_country(seller_name, seller_address, currency)



    return {
        "公司名称": seller_name if seller_name else "未识别",
        "国内/海外": country,
        "货币类型": currency if currency else "未识别",
        "金额": amount if amount is not None else "未识别",
        "支付日期": payment_date,
        "项目名称": project_name,
        "文件名": filename,
        "备注": ""
    }


def sanitize_filename_part(value):
    if value is None:
        return "unknown"
    text = str(value).strip()
    if not text or text == "未识别":
        return "unknown"
    text = re.sub(r'[<>:"/\\|?*]+', "_", text)
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"_+", "_", text).strip("._ ")
    return text or "unknown"


def simplify_company_for_filename(company_name):
    if not company_name:
        return "unknown"
    text = str(company_name).strip()
    if not text or text == "未识别":
        return "unknown"

    normalized_for_alias = re.sub(r"[^a-z0-9\s]+", " ", text.lower())
    normalized_for_alias = re.sub(r"\s+", " ", normalized_for_alias).strip()
    if normalized_for_alias in COMPANY_SHORT_NAME_ALIASES:
        return COMPANY_SHORT_NAME_ALIASES[normalized_for_alias]

    # Remove common legal suffixes, keep the core brand name.
    text = re.sub(
        r"\b(co\.?,?\s*ltd\.?|co\.?,?\s*limited|limited|ltd\.?|inc\.?|llc|company|corp\.?|corporation)\b",
        "",
        text,
        flags=re.I,
    )
    text = re.sub(r"\s+", " ", text).strip(" ,.-_")
    if not text:
        return "unknown"

    # Try alias mapping again after suffix cleanup.
    normalized_after_cleanup = re.sub(r"[^a-z0-9\s]+", " ", text.lower())
    normalized_after_cleanup = re.sub(r"\s+", " ", normalized_after_cleanup).strip()
    if normalized_after_cleanup in COMPANY_SHORT_NAME_ALIASES:
        return COMPANY_SHORT_NAME_ALIASES[normalized_after_cleanup]

    # Keep first token for concise filename, e.g. "Comets International" -> "Comets".
    return text.split()[0]


def build_invoice_filename(fields):
    payment_date = fields.get("支付日期", "")
    year_month = datetime.now().strftime("%Y%m")
    if isinstance(payment_date, str):
        match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", payment_date)
        if match:
            year_month = f"{match.group(1)}{int(match.group(2)):02d}"

    short_company = simplify_company_for_filename(fields.get("公司名称"))
    company = sanitize_filename_part(short_company)
    project = sanitize_filename_part(fields.get("项目名称"))
    amount = fields.get("金额")
    amount_text = canonicalize_amount_for_filename(amount)

    return f"{year_month}_{company}_{project}_{amount_text}.pdf"


def should_skip_ocr_rename(file_name):
    return bool(re.match(r"^\d{6}_.+_.+_.+\.pdf$", file_name, re.I))


def canonicalize_amount_for_filename(amount):
    if amount is None or amount == "" or amount == "未识别":
        return "unknown"
    if isinstance(amount, (int, float)):
        text = f"{amount:.2f}"
    else:
        text = str(amount).strip()
    # Normalize: keep digits + optional dot, strip commas/currency.
    text = re.sub(r"[^0-9.]+", "", text.replace(",", ""))
    if not text:
        return "unknown"
    # Trim trailing zeros (4367.40 -> 4367.4; 10000.00 -> 10000)
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return sanitize_filename_part(text)


def filename_matches_fields(file_name, fields):
    """
    Validate whether a filename matches extracted (scanned) fields.
    If it doesn't match, we will rename it again to the canonical name.
    """
    m = re.match(r"^(?P<ym>\d{6})_(?P<company>.+)_(?P<project>.+)_(?P<amount>.+)\.pdf$", file_name, re.I)
    if not m:
        return False

    expected = build_invoice_filename(fields)
    m2 = re.match(r"^(?P<ym>\d{6})_(?P<company>.+)_(?P<project>.+)_(?P<amount>.+)\.pdf$", expected, re.I)
    if not m2:
        return False

    def _norm(s):
        return re.sub(r"[_\s]+", "", (s or "")).lower()

    # Compare canonical parts; allow minor formatting differences in amount.
    if m.group("ym") != m2.group("ym"):
        return False
    if _norm(m.group("company")) != _norm(m2.group("company")):
        return False
    if _norm(m.group("project")) != _norm(m2.group("project")):
        return False

    amt_a = canonicalize_amount_for_filename(m.group("amount"))
    amt_b = canonicalize_amount_for_filename(fields.get("金额"))
    return _norm(amt_a) == _norm(amt_b)


def rename_invoice_file_if_needed(old_path, old_name, fields):
    # If it's already in our target pattern AND matches extracted fields, skip.
    if should_skip_ocr_rename(old_name) and filename_matches_fields(old_name, fields):
        return old_name

    new_name = build_invoice_filename(fields)
    if new_name.lower() == old_name.lower():
        return old_name

    parent_dir = os.path.dirname(old_path)
    base, ext = os.path.splitext(new_name)
    candidate = new_name
    idx = 1
    while os.path.exists(os.path.join(parent_dir, candidate)):
        candidate = f"{base}_{idx}{ext}"
        idx += 1

    new_path = os.path.join(parent_dir, candidate)
    os.rename(old_path, new_path)
    logging.info(f"[rename] {old_name} -> {candidate}")
    return candidate


def archive_processed_file(file_name):
    """
    Move processed invoice into a date-labeled archive path:
    archive/YYYYMMDD/
    """
    source_path = os.path.join(input_folder, file_name)
    if not os.path.exists(source_path):
        logging.warning(f"[archive] Source file not found, skipping archive: {file_name}")
        return ""

    run_date = datetime.now()
    ymd = run_date.strftime("%Y%m%d")
    archive_dir = os.path.join(archive_root_folder, ymd)
    os.makedirs(archive_dir, exist_ok=True)

    base, ext = os.path.splitext(file_name)
    candidate = file_name
    idx = 1
    while os.path.exists(os.path.join(archive_dir, candidate)):
        candidate = f"{base}_{idx}{ext}"
        idx += 1

    target_path = os.path.join(archive_dir, candidate)
    shutil.move(source_path, target_path)
    logging.info(f"[archive] Moved to archive: {candidate} -> {archive_dir}")
    return target_path


def clear_input_folder():
    """Clear remaining files/subfolders in input_folder after a batch run."""
    if not os.path.exists(input_folder):
        return

    for name in os.listdir(input_folder):
        path = os.path.join(input_folder, name)
        try:
            if os.path.isfile(path):
                os.remove(path)
            elif os.path.isdir(path):
                shutil.rmtree(path)
        except Exception as e:
            logging.warning(f"[cleanup] Failed to delete: {path}, error: {e}")

    logging.info("[cleanup] input_folder cleared")


def format_excel_output(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    # Header color: rgb(172, 185, 202) => FFACB9CA (ARGB)
    header_fill = PatternFill(fill_type="solid", start_color="FFACB9CA", end_color="FFACB9CA")
    header_font = Font(name="DengXian", size=12, color="FF000000")
    body_font = Font(name="DengXian", size=12, color="FF000000")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    ws.row_dimensions[1].height = ws.sheet_format.defaultRowHeight

    # Style body rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = body_font
            cell.alignment = body_alignment

    # Auto-size columns with extra padding for readability
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        # Add padding and keep a sensible min/max width
        ws.column_dimensions[col_letter].width = min(max(16, max_len + 4), 60)

    wb.save(excel_path)


def run_invoice_processing(
    open_excel_after: bool = False,
    progress_callback: Optional[Callable[[int, int, str], None]] = None,
):
    """
    Process all PDFs in input_folder and generate an Excel output.
    Returns: {"rows": [...], "output_file": str|None, "processed_files": int}

    Uses pdfplumber when possible, else Poppler→Tesseract; fields via regex/heuristics.
    """
    if not os.path.exists(input_folder):
        os.makedirs(input_folder)
        logging.info("Created input folder. Add PDF files and run again.")
        return {"rows": [], "output_file": None, "processed_files": 0}

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    check_poppler()

    pdf_files = [f for f in os.listdir(input_folder) if f.lower().endswith(".pdf")]
    logging.info(f"Found files: {pdf_files}")
    total_files = len(pdf_files)

    rows = []

    for idx, file in enumerate(pdf_files, start=1):
        pdf_path = os.path.join(input_folder, file)
        if progress_callback:
            progress_callback(idx, total_files, f"Processing ({idx}/{total_files}): {file}")
        logging.info(f"Processing file: {file}")

        text, _used_ocr = extract_pdf_text(pdf_path)
        logging.info(f"[debug][{file}] contains '星痕共鸣': {'星痕共鸣' in text}")
        if not text:
            logging.warning(f"Could not extract text: {file}")
            rows.append({
                "公司名称": "OCR失败",
                "国内/海外": "",
                "货币类型": "",
                "金额": "",
                "支付日期": "",
                "项目名称": "",
                "文件名": file,
                "备注": ""
            })
            continue

        fields = extract_fields(text, file)
        final_file_name = rename_invoice_file_if_needed(pdf_path, file, fields)
        fields["文件名"] = final_file_name
        archive_processed_file(final_file_name)
        rows.append(fields)
        if file == debug_target_pdf:
            logging.info(f"[debug][{file}] extracted fields: {fields}")

    output_path = None
    if rows:
        today_local = datetime.now().strftime("%Y%m%d")
        output_path = os.path.join(output_folder, f"invoice_result_{today_local}.xlsx")
        df = pd.DataFrame(rows)
        df.to_excel(output_path, index=False)
        format_excel_output(output_path)
        logging.info(f"PDF processing completed. Output file: {output_path}")
        if progress_callback:
            progress_callback(total_files, total_files, "Done. Excel file generated.")
    else:
        logging.warning("No PDF files were processed successfully.")
        if progress_callback:
            progress_callback(total_files, total_files, "No PDFs were processed successfully.")

    if open_excel_after and output_path:
        try:
            os.startfile(output_path)
            logging.info(f"Opened Excel file automatically: {output_path}")
        except Exception as e:
            logging.error(f"Failed to auto-open Excel file: {e}")

    clear_input_folder()

    return {
        "rows": rows,
        "output_file": output_path,
        "processed_files": total_files,
    }